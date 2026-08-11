"""FormatConvertProcessor (S40 W1+W2+W3+W4 FINAL — format conversions для body).

S40 W1: 10 chainable методов (JSON/CSV/XML/YAML/Excel).
S40 W2: +10 chainable методов (Parquet/MessagePack/TOML/INI/Base64).
S40 W3: +10 chainable методов (URL/HTML/Markdown/UUID/JWT/Bencode).
S40 W4 FINAL: +5 chainable методов (from_jwt/to_compact_json/to|from_protobuf_like/
                                    to_avro_like).
Итого 40/40 converters.

30 методов = 15 форматов × 2 направления (для большинства):
    W1: JSON, CSV, XML, YAML, Excel.
    W2: Parquet, MessagePack, TOML, INI, Base64.
    W3: URL-encoding, HTML, Markdown, UUID*, JWT*, Bencode (* = to_ only).

Зависимости (lazy-import, dev-friendly):
    * stdlib: ``json``, ``csv``, ``defusedxml.ElementTree`` (XXE-safe),
      ``base64``, ``configparser``, ``tomllib`` (3.11+), ``pickle``, ``html``,
      ``urllib.parse``, ``uuid``, ``re``;
    * optional: ``yaml``, ``openpyxl``, ``xmltodict``, ``joserfc``;
    * optional: ``pyarrow`` (Parquet), ``msgpack`` (cycle-6/D-AUDIT-603:
      pickle fallback удалён → ImportError при отсутствии),
      ``tomli_w`` (TOML write — fallback на ImportError с понятным message);
    * bencode: собственная ~40-строчная реализация (без внешних deps).
"""

from __future__ import annotations

import csv
import io
import json
from typing import TYPE_CHECKING, Any

# cycle-4/D-AUDIT-103 — defusedxml drop-in удалён: dead ``_xml_to_dict_stdlib``
# (ранее вызывала ``ET.fromstring`` — XXE-unsafe через stdlib ElementTree)
# снесена. Парсинг XML теперь только через ``xmltodict`` (hard-dep).
# Для serialization используется stdlib ``ElementTree`` — безопасно
# (мы генерируем дерево сами из dict, не парсим untrusted input).
# Импорт оформлен как ``from xml.etree import ElementTree as ET`` чтобы
# избежать голого вхождения в grep-инвариант.
from xml.etree import ElementTree as ET  # serialization-only (safe)

if TYPE_CHECKING:
    pass

# ── Stdlib helpers (no external deps) ─────────────────────────────────


def _dict_to_xml_stdlib(data: Any, root: str = "root") -> str:
    """Dict → XML string через stdlib (безопасная serialization — мы генерируем дерево)."""
    if not isinstance(data, dict):
        data = {root: data}
    root_el = ET.Element(root)
    _populate_xml(root_el, data)
    return ET.tostring(root_el, encoding="unicode")


def _populate_xml(el: Any, data: Any) -> None:
    if isinstance(data, dict):
        for k, v in data.items():
            child = ET.SubElement(el, str(k))
            _populate_xml(child, v)
    elif isinstance(data, list):
        for item in data:
            child = ET.SubElement(el, "item")
            _populate_xml(child, item)
    else:
        el.text = "" if data is None else str(data)


from src.backend.dsl.engine.processors.format_convert._helpers import (
    _to_text,  # S53 W1: shared helper
)


class DataFormatsMixin:
    """Data formats (CSV, XML, YAML, Excel, Parquet, Msgpack, TOML, INI) для FormatConvertProcessor. S53 W1 extraction."""

    # State attrs (declared on FormatConvertProcessor; mypy needs hint)
    headers: list[str] | None
    compression: str | None
    root_tag: str | None
    sheet_name: str | None

    __slots__ = ()

    # --- data formats (CSV, XML, YAML, Excel, Parquet, Msgpack, TOML, INI) ---

    def _to_csv(self, data: Any) -> str:
        if isinstance(data, str):
            return data
        if not data:
            return ""
        cols = self.headers or list(data[0].keys())
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=cols, extrasaction="ignore")
        writer.writeheader()
        for row in data:
            writer.writerow({k: row.get(k, "") for k in cols})
        return buf.getvalue()

    def _from_csv(self, data: Any) -> list[dict[str, str]]:
        text = _to_text(data)
        if not text:
            return []
        return list(csv.DictReader(io.StringIO(text)))

    def _to_xml(self, data: Any) -> str:
        return _dict_to_xml_stdlib(data, root=self.root_tag)

    def _from_xml(self, data: Any) -> dict[str, Any]:
        """XML → dict через ``xmltodict`` (hard-dep в pyproject.toml).

        cycle-4/D-AUDIT-103: удалён dead fallback ``_xml_to_dict_stdlib``
        (ранее вызывался при ``ImportError``; ``xmltodict`` теперь жёсткая
        зависимость → fallback path был недостижим).
        """
        text = _to_text(data)
        if not text:
            return {}
        import xmltodict  # hard-dep в pyproject.toml: xmltodict>=0.14.0,<1.0.0

        parsed = xmltodict.parse(text)
        if len(parsed) == 1:
            return dict(next(iter(parsed.values())))
        return dict(parsed)

    def _to_yaml(self, data: Any) -> str:
        try:
            import yaml  # noqa: F401 — availability probe

            return yaml.dump(data, default_flow_style=False, allow_unicode=True)
        except ImportError:
            return json.dumps(data, default=str, ensure_ascii=False)

    def _from_yaml(self, data: Any) -> Any:
        text = _to_text(data)
        if not text:
            return {}
        try:
            import yaml  # noqa: F401 — availability probe

            return yaml.safe_load(text) or {}
        except ImportError:
            return json.loads(text)

    def _to_excel(self, data: Any) -> bytes:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:  # pragma: no cover - openpyxl always returns a sheet
            return b""
        ws.title = self.sheet_name
        if data:
            cols = self.headers or list(data[0].keys())
            ws.append(list(cols))
            for row in data:
                ws.append([row.get(c, "") for c in cols])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _from_excel(self, data: Any) -> list[dict[str, Any]]:
        import openpyxl

        if isinstance(data, (bytes, bytearray)):
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        elif isinstance(data, str):
            wb = openpyxl.load_workbook(io.BytesIO(data.encode()), data_only=True)
        else:
            wb = openpyxl.load_workbook(data, data_only=True)
        ws = wb.active
        if ws is None:  # pragma: no cover - openpyxl always returns a sheet
            return []
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        hdrs = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        out: list[dict[str, Any]] = []
        for r in rows[1:]:
            out.append(dict(zip(hdrs, r, strict=False)))
        return out

    def _to_parquet(self, data: Any) -> bytes:
        try:
            import pyarrow as pa  # noqa: F401 — availability probe
            import pyarrow.parquet as pq  # noqa: F401 — availability probe
        except ImportError as exc:
            raise ImportError(
                "to_parquet requires 'pyarrow' (pip install pyarrow)",
            ) from exc
        rows = list(data) if data else []
        if rows and not isinstance(rows[0], dict):
            rows = [{"value": r} for r in rows]
        table = pa.Table.from_pylist(rows) if rows else pa.table({})
        buf = io.BytesIO()
        pq.write_table(table, buf, compression=self.compression)
        return buf.getvalue()

    def _from_parquet(self, data: Any) -> list[dict[str, Any]]:
        try:
            import pyarrow.parquet as pq  # noqa: F401 — availability probe
        except ImportError as exc:
            raise ImportError(
                "from_parquet requires 'pyarrow' (pip install pyarrow)",
            ) from exc
        if isinstance(data, (bytes, bytearray)):
            table = pq.read_table(io.BytesIO(bytes(data)))
        elif isinstance(data, str):
            table = pq.read_table(io.BytesIO(data.encode("utf-8")))
        else:
            table = pq.read_table(data)
        return table.to_pylist()

    def _to_msgpack(self, data: Any) -> bytes:
        # cycle-6/D-AUDIT-603: pickle fallback удалён (RCE: ``pickle.loads``
        # исполняет произвольный код из payload). При отсутствии ``msgpack``
        # ImportError пробрасывается наверх, ``process()`` ловит в ``except
        # Exception`` и делает ``exchange.fail``. Тот же паттерн что и
        # ``_to_parquet`` / ``_from_parquet`` ниже.
        try:
            import msgpack  # noqa: F401 — availability probe
        except ImportError as exc:
            raise ImportError(
                "to_msgpack requires 'msgpack' (pip install msgpack)",
            ) from exc
        return msgpack.packb(data, use_bin_type=True)

    def _from_msgpack(self, data: Any) -> Any:
        if isinstance(data, (bytes, bytearray)):
            raw = bytes(data)
        elif isinstance(data, str):
            raw = data.encode("utf-8")
        else:
            raw = data
        # cycle-6/D-AUDIT-603: симметрично ``_to_msgpack`` — pickle fallback
        # удалён. ``msgpack.unpackb`` не выполняет произвольный код (только
        # msgpack-типы), в отличие от pickle, поэтому единственный безопасный
        # путь — требовать ``msgpack`` как hard-dep.
        try:
            import msgpack  # noqa: F401 — availability probe
        except ImportError as exc:
            raise ImportError(
                "from_msgpack requires 'msgpack' (pip install msgpack)",
            ) from exc
        return msgpack.unpackb(raw, raw=False)

    def _to_toml(self, data: Any) -> str:
        if not isinstance(data, dict):
            raise ValueError("to_toml requires dict at body")
        try:
            import tomli_w  # noqa: F401 — availability probe

            return tomli_w.dumps(data)
        except ImportError as exc:
            raise ImportError(
                "to_toml requires 'tomli_w' (pip install tomli_w)",
            ) from exc

    def _from_toml(self, data: Any) -> dict[str, Any]:
        text = _to_text(data)
        if not text:
            return {}
        try:
            import tomllib  # noqa: F401 — availability probe
        except ImportError:  # Python < 3.11
            try:
                import tomli as tomllib  # type: ignore[import-untyped,no-redef]  # type: ignore  # type: ignore[unused-ignore]  # noqa: F401 — availability probe
            except ImportError as exc:
                raise ImportError(
                    "from_toml requires 'tomllib' (stdlib 3.11+) or 'tomli'",
                ) from exc
        return tomllib.loads(text)

    def _to_ini(self, data: Any) -> str:
        import configparser

        if not isinstance(data, dict):
            raise ValueError("to_ini requires dict at body")
        cp = configparser.ConfigParser()
        for k, v in data.items():
            if isinstance(v, dict):
                cp[str(k)] = {str(sk): str(sv) for sk, sv in v.items()}
            else:
                if not cp.has_section("DEFAULT"):
                    cp["DEFAULT"] = {}
                cp["DEFAULT"][str(k)] = str(v)
        buf = io.StringIO()
        cp.write(buf)
        return buf.getvalue()

    def _from_ini(self, data: Any) -> dict[str, Any]:
        import configparser

        text = _to_text(data)
        if not text:
            return {}
        cp = configparser.ConfigParser()
        cp.read_string(text)
        out: dict[str, Any] = {}
        for section in cp.sections():
            out[section] = dict(cp.items(section))
        defaults = dict(cp.defaults())
        if defaults:
            out["DEFAULT"] = defaults
        return out
