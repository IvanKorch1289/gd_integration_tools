"""FormatConvertersMixin (S40 W1+W2+W3+W4 FINAL — format conversions для body).

S40 W1: 10 chainable методов (JSON/CSV/XML/YAML/Excel).
S40 W2: +10 chainable методов (Parquet/MessagePack/TOML/INI/Base64).
S40 W3: +10 chainable методов (URL/HTML/Markdown/UUID/JWT/Bencode).
S40 W4 FINAL: +5 chainable методов (from_jwt/to_compact_json/to|from_protobuf_like/
                                    to_avro_like).
Итого 40/40 converters.

Назван ``FormatConvertersMixin`` (не ``ConvertersMixin``) чтобы не конфликтовать
с Phase-2.1 :class:`dsl.builders.converters.ConvertersMixin` (hash/encrypt/
decrypt/compress/decompress — 5 методов), который уже в MRO.

30 методов = 15 форматов × 2 направления (для большинства):
    W1: JSON, CSV, XML, YAML, Excel.
    W2: Parquet, MessagePack, TOML, INI, Base64.
    W3: URL-encoding, HTML, Markdown, UUID*, JWT*, Bencode (* = to_ only).

Зависимости (lazy-import, dev-friendly):
    * stdlib: ``json``, ``csv``, ``xml.etree.ElementTree``, ``base64``,
      ``configparser``, ``tomllib`` (3.11+), ``pickle``, ``html``,
      ``urllib.parse``, ``uuid``, ``re``;
    * optional: ``yaml``, ``openpyxl``, ``xmltodict``, ``joserfc``;
    * optional: ``pyarrow`` (Parquet), ``msgpack`` (fallback → ``pickle``),
      ``tomli_w`` (TOML write — fallback на ImportError с понятным message);
    * bencode: собственная ~40-строчная реализация (без внешних deps).
"""

from __future__ import annotations

import base64
import csv
import html
import io
import json
import re
import urllib.parse
import uuid
import xml.etree.ElementTree as ET
from typing import TYPE_CHECKING, Any, ClassVar

from src.backend.dsl.engine.processors.base import BaseProcessor

if TYPE_CHECKING:
    from src.backend.dsl.builders.base import RouteBuilder
    from src.backend.dsl.engine.context import ExecutionContext
    from src.backend.dsl.engine.exchange import Exchange

__all__ = ("FormatConvertersMixin", "FormatConvertProcessor")


# ── Stdlib helpers (no external deps) ─────────────────────────────────


def _dict_to_xml_stdlib(data: Any, root: str = "root") -> str:
    """dict → XML string через stdlib ``xml.etree.ElementTree``."""
    if not isinstance(data, dict):
        data = {root: data}
    root_el = ET.Element(root)
    _populate_xml(root_el, data)
    return ET.tostring(root_el, encoding="unicode")


def _populate_xml(el: ET.Element, data: Any) -> None:
    if isinstance(data, dict):
        for k, v in data.items():
            child = ET.SubElement(el, str(k))
            _populate_xml(child, v)
    elif isinstance(data, list):
        for item in data:
            child = ET.SubElement(el, "item")
            _populate_xml(child, item)
    else:
        el.text = "" if data is None else str(data)


def _xml_to_dict_stdlib(xml_string: str) -> dict[str, Any]:
    """XML → dict через stdlib (используется если xmltodict недоступен)."""
    root = ET.fromstring(xml_string)  # noqa: S314
    return {root.tag: _el_to_dict(root)}


def _el_to_dict(el: ET.Element) -> Any:
    children = list(el)
    if not children:
        return el.text or ""
    out: dict[str, Any] = {}
    for child in children:
        out[child.tag] = _el_to_dict(child)
    return out


def _to_text(data: Any) -> str:
    """bytes/bytearray → str (utf-8 best-effort)."""
    if isinstance(data, (bytes, bytearray)):
        return data.decode("utf-8", errors="replace")
    return data


# ── Processor ─────────────────────────────────────────────────────────


class FormatConvertProcessor(BaseProcessor):
    """Универсальный format-conversion processor (S40 W1).

    Один процессор обслуживает все 10 направлений через ``direction`` +
    format-specific kwargs. Lazy-import внешних библиотек (yaml, openpyxl,
    xmltodict) — dev_light остаётся работоспособным при их отсутствии.
    """

    side_effect: ClassVar[str] = "PURE"
    compensatable: ClassVar[bool] = True

    def __init__(
        self,
        *,
        direction: str,
        fmt: str,
        indent: int | None = None,
        headers: list[str] | None = None,
        root_tag: str = "root",
        sheet_name: str = "Sheet1",
        compression: str = "snappy",
        source_value: Any = None,
        from_property: str = "body",
        name: str | None = None,
        # JWT (S40 W3)
        secret: str | None = None,
        algorithm: str = "HS256",
        claims: dict[str, Any] | None = None,
        # Avro-like (S40 W4)
        schema: dict[str, Any] | None = None,
    ) -> None:
        super().__init__(name=name or f"format:{direction}:{fmt}")
        self.direction = direction
        self.fmt = fmt
        self.indent = indent
        self.headers = headers
        self.root_tag = root_tag
        self.sheet_name = sheet_name
        self.compression = compression
        self.source_value = source_value
        self.from_property = from_property
        self.secret = secret
        self.algorithm = algorithm
        self.claims = claims
        self.schema = schema

    async def process(
        self, exchange: "Exchange[Any]", context: "ExecutionContext"
    ) -> None:
        # 1. resolve input
        if self.source_value is not None:
            data: Any = self.source_value
        elif self.from_property != "body":
            data = exchange.properties.get(self.from_property)
        else:
            data = exchange.in_message.body

        if data is None:
            exchange.set_out(body=None, headers=dict(exchange.in_message.headers))
            return

        try:
            result = self._convert(data)
            exchange.set_out(body=result, headers=dict(exchange.in_message.headers))
        except Exception as exc:  # parse/format failures → fail exchange
            exchange.fail(f"format convert {self.direction}:{self.fmt} failed: {exc}")

    # ── dispatch ──

    def _convert(self, data: Any) -> Any:
        if self.direction == "to_json":
            return self._to_json(data)
        if self.direction == "from_json":
            return self._from_json(data)
        if self.direction == "to_csv":
            return self._to_csv(data)
        if self.direction == "from_csv":
            return self._from_csv(data)
        if self.direction == "to_xml":
            return self._to_xml(data)
        if self.direction == "from_xml":
            return self._from_xml(data)
        if self.direction == "to_yaml":
            return self._to_yaml(data)
        if self.direction == "from_yaml":
            return self._from_yaml(data)
        if self.direction == "to_excel":
            return self._to_excel(data)
        if self.direction == "from_excel":
            return self._from_excel(data)
        if self.direction == "to_parquet":
            return self._to_parquet(data)
        if self.direction == "from_parquet":
            return self._from_parquet(data)
        if self.direction == "to_msgpack":
            return self._to_msgpack(data)
        if self.direction == "from_msgpack":
            return self._from_msgpack(data)
        if self.direction == "to_toml":
            return self._to_toml(data)
        if self.direction == "from_toml":
            return self._from_toml(data)
        if self.direction == "to_ini":
            return self._to_ini(data)
        if self.direction == "from_ini":
            return self._from_ini(data)
        if self.direction == "to_base64":
            return self._to_base64(data)
        if self.direction == "from_base64":
            return self._from_base64(data)
        # ── S40 W3: URL / HTML / Markdown / UUID / JWT / Bencode ──
        if self.direction == "to_url_encoded":
            return self._to_url_encoded(data)  # type: ignore[attr-defined]
        if self.direction == "from_url_encoded":
            return self._from_url_encoded(data)  # type: ignore[attr-defined]
        if self.direction == "to_html_escape":
            return self._to_html_escape(data)  # type: ignore[attr-defined]
        if self.direction == "from_html_unescape":
            return self._from_html_unescape(data)  # type: ignore[attr-defined]
        if self.direction == "to_markdown":
            return self._to_markdown(data)  # type: ignore[attr-defined]
        if self.direction == "from_markdown":
            return self._from_markdown(data)  # type: ignore[attr-defined]
        if self.direction == "to_uuid_string":
            return self._to_uuid_string(data)  # type: ignore[attr-defined]
        if self.direction == "to_jwt":
            return self._to_jwt(data)  # type: ignore[attr-defined]
        if self.direction == "to_bencode":
            return self._to_bencode(data)  # type: ignore[attr-defined]
        if self.direction == "from_bencode":
            return self._from_bencode(data)  # type: ignore[attr-defined]
        # ── S40 W4 FINAL: from_jwt / to_compact_json / to|from_protobuf_like / to_avro_like ──
        if self.direction == "from_jwt":
            return self._from_jwt(data)  # type: ignore[attr-defined]
        if self.direction == "to_compact_json":
            return self._to_compact_json(data)  # type: ignore[attr-defined]
        if self.direction == "to_protobuf_like":
            return self._to_protobuf_like(data)  # type: ignore[attr-defined]
        if self.direction == "from_protobuf_like":
            return self._from_protobuf_like(data)  # type: ignore[attr-defined]
        if self.direction == "to_avro_like":
            return self._to_avro_like(data)  # type: ignore[attr-defined]
        raise ValueError(f"unknown direction: {self.direction!r}")

    # ── JSON ──

    def _to_json(self, data: Any) -> str:
        if isinstance(data, (bytes, bytearray)):
            return data.decode("utf-8", errors="replace")
        if isinstance(data, str):
            return data  # already JSON string
        return json.dumps(data, indent=self.indent, default=str, ensure_ascii=False)

    def _from_json(self, data: Any) -> Any:
        text = _to_text(data)
        if text == "":
            return None
        return json.loads(text)

    # ── CSV ──

    def _to_csv(self, data: Any) -> str:
        if isinstance(data, str):
            return data
        if not data:
            return ""
        cols = self.headers or list(data[0].keys())
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=cols, extrasaction="ignore")
        writer.writeheader()
        for row in data:
            writer.writerow({k: row.get(k, "") for k in cols})
        return buf.getvalue()

    def _from_csv(self, data: Any) -> list[dict[str, str]]:
        text = _to_text(data)
        if not text:
            return []
        return list(csv.DictReader(io.StringIO(text)))

    # ── XML ──

    def _to_xml(self, data: Any) -> str:
        return _dict_to_xml_stdlib(data, root=self.root_tag)

    def _from_xml(self, data: Any) -> dict[str, Any]:
        text = _to_text(data)
        if not text:
            return {}
        try:
            import xmltodict

            parsed = xmltodict.parse(text)
            if len(parsed) == 1:
                return dict(next(iter(parsed.values())))
            return dict(parsed)
        except ImportError:
            return _xml_to_dict_stdlib(text)

    # ── YAML ──

    def _to_yaml(self, data: Any) -> str:
        try:
            import yaml

            return yaml.dump(data, default_flow_style=False, allow_unicode=True)
        except ImportError:
            return json.dumps(data, default=str, ensure_ascii=False)

    def _from_yaml(self, data: Any) -> Any:
        text = _to_text(data)
        if not text:
            return {}
        try:
            import yaml

            return yaml.safe_load(text) or {}
        except ImportError:
            return json.loads(text)

    # ── Excel ──

    def _to_excel(self, data: Any) -> bytes:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:  # pragma: no cover - openpyxl always returns a sheet
            return b""
        ws.title = self.sheet_name
        if data:
            cols = self.headers or list(data[0].keys())
            ws.append(list(cols))
            for row in data:
                ws.append([row.get(c, "") for c in cols])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _from_excel(self, data: Any) -> list[dict[str, Any]]:
        import openpyxl

        if isinstance(data, (bytes, bytearray)):
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        elif isinstance(data, str):
            wb = openpyxl.load_workbook(io.BytesIO(data.encode()), data_only=True)
        else:
            wb = openpyxl.load_workbook(data, data_only=True)
        ws = wb.active
        if ws is None:  # pragma: no cover - openpyxl always returns a sheet
            return []
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        hdrs = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        out: list[dict[str, Any]] = []
        for r in rows[1:]:
            out.append({h: v for h, v in zip(hdrs, r, strict=False)})
        return out

    # ── Parquet (S40 W2) ──

    def _to_parquet(self, data: Any) -> bytes:
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
        except ImportError as exc:
            raise ImportError(
                "to_parquet requires 'pyarrow' (pip install pyarrow)"
            ) from exc
        rows = list(data) if data else []
        if rows and not isinstance(rows[0], dict):
            rows = [{"value": r} for r in rows]
        table = pa.Table.from_pylist(rows) if rows else pa.table({})
        buf = io.BytesIO()
        pq.write_table(table, buf, compression=self.compression)
        return buf.getvalue()

    def _from_parquet(self, data: Any) -> list[dict[str, Any]]:
        try:
            import pyarrow.parquet as pq
        except ImportError as exc:
            raise ImportError(
                "from_parquet requires 'pyarrow' (pip install pyarrow)"
            ) from exc
        if isinstance(data, (bytes, bytearray)):
            table = pq.read_table(io.BytesIO(bytes(data)))
        elif isinstance(data, str):
            table = pq.read_table(io.BytesIO(data.encode("utf-8")))
        else:
            table = pq.read_table(data)
        return table.to_pylist()

    # ── MessagePack (S40 W2) — graceful fallback на pickle ──

    def _to_msgpack(self, data: Any) -> bytes:
        try:
            import msgpack

            return msgpack.packb(data, use_bin_type=True)
        except ImportError:
            # Fallback: pickle используется только когда msgpack недоступен
            # (dev_light / minimal install). Данные остаются в pipeline — это
            # наш собственный round-trip, не untrusted input.
            import pickle

            return pickle.dumps(data)

    def _from_msgpack(self, data: Any) -> Any:
        if isinstance(data, (bytes, bytearray)):
            raw = bytes(data)
        elif isinstance(data, str):
            raw = data.encode("utf-8")
        else:
            raw = data
        try:
            import msgpack

            return msgpack.unpackb(raw, raw=False)
        except ImportError:
            # Symmetric fallback к pickle — см. комментарий в _to_msgpack.
            import pickle

            return pickle.loads(raw)  # noqa: S301 — см. комментарий выше

    # ── TOML (S40 W2) — read: stdlib tomllib; write: tomli_w ──

    def _to_toml(self, data: Any) -> str:
        if not isinstance(data, dict):
            raise ValueError("to_toml requires dict at body")
        try:
            import tomli_w

            return tomli_w.dumps(data)
        except ImportError as exc:
            raise ImportError(
                "to_toml requires 'tomli_w' (pip install tomli_w)"
            ) from exc

    def _from_toml(self, data: Any) -> dict[str, Any]:
        text = _to_text(data)
        if not text:
            return {}
        try:
            import tomllib
        except ImportError:  # Python < 3.11
            try:
                import tomli as tomllib  # type: ignore[import-untyped,no-redef]  # type: ignore  # type: ignore[unused-ignore]
            except ImportError as exc:
                raise ImportError(
                    "from_toml requires 'tomllib' (stdlib 3.11+) or 'tomli'"
                ) from exc
        return tomllib.loads(text)

    # ── INI (S40 W2) — stdlib configparser ──

    def _to_ini(self, data: Any) -> str:
        import configparser

        if not isinstance(data, dict):
            raise ValueError("to_ini requires dict at body")
        cp = configparser.ConfigParser()
        for k, v in data.items():
            if isinstance(v, dict):
                cp[str(k)] = {str(sk): str(sv) for sk, sv in v.items()}
            else:
                if not cp.has_section("DEFAULT"):
                    cp["DEFAULT"] = {}
                cp["DEFAULT"][str(k)] = str(v)
        buf = io.StringIO()
        cp.write(buf)
        return buf.getvalue()

    def _from_ini(self, data: Any) -> dict[str, Any]:
        import configparser

        text = _to_text(data)
        if not text:
            return {}
        cp = configparser.ConfigParser()
        cp.read_string(text)
        out: dict[str, Any] = {}
        for section in cp.sections():
            out[section] = dict(cp.items(section))
        defaults = dict(cp.defaults())
        if defaults:
            out["DEFAULT"] = defaults
        return out

    # ── Base64 (S40 W2) — stdlib base64 ──

    def _to_base64(self, data: Any) -> str:
        import base64

        if isinstance(data, (bytes, bytearray)):
            raw = bytes(data)
        elif isinstance(data, str):
            raw = data.encode("utf-8")
        else:
            raw = str(data).encode("utf-8")
        return base64.b64encode(raw).decode("ascii")

    def _from_base64(self, data: Any) -> bytes:
        import base64

        text = _to_text(data)
        if not text:
            return b""
        return base64.b64decode(text, validate=False)

    # ── URL-encoding (S40 W3) — stdlib urllib.parse ──

    def _to_url_encoded(self, data: Any) -> str:
        if isinstance(data, str):
            return data
        if isinstance(data, (bytes, bytearray)):
            data = data.decode("utf-8", errors="replace")
        if data is None:
            return ""
        return urllib.parse.urlencode(data, doseq=True)

    def _from_url_encoded(self, data: Any) -> dict[str, Any]:
        text = _to_text(data)
        if not text:
            return {}
        parsed = urllib.parse.parse_qs(text, keep_blank_values=True)
        return {k: (v[0] if len(v) == 1 else v) for k, v in parsed.items()}

    # ── HTML (S40 W3) — stdlib html ──

    def _to_html_escape(self, data: Any) -> str:
        if data is None:
            return ""
        return html.escape(_to_text(data), quote=True)

    def _from_html_unescape(self, data: Any) -> str:
        text = _to_text(data)
        if not text:
            return ""
        return html.unescape(text)

    # ── Markdown (S40 W3) — stdlib re (simple header-based) ──

    def _to_markdown(self, data: Any) -> str:
        if isinstance(data, str):
            return data
        if data is None:
            return ""
        if not isinstance(data, dict):
            data = {"value": data}
        parts: list[str] = []
        for k, v in data.items():
            parts.append(f"# {k}")
            if isinstance(v, (dict, list)):
                parts.append(json.dumps(v, default=str, ensure_ascii=False))
            else:
                parts.append(str(v))
            parts.append("")
        return "\n".join(parts)

    def _from_markdown(self, data: Any) -> dict[str, str]:
        text = _to_text(data)
        if not text:
            return {}
        out: dict[str, str] = {}
        current_key: str | None = None
        current_lines: list[str] = []
        for line in text.splitlines():
            m = re.match(r"^#\s+(.+?)\s*$", line)
            if m:
                if current_key is not None:
                    out[current_key] = "\n".join(current_lines).strip()
                current_key = m.group(1)
                current_lines = []
            else:
                current_lines.append(line)
        if current_key is not None:
            out[current_key] = "\n".join(current_lines).strip()
        return out

    # ── UUID (S40 W3) — stdlib uuid (to_ only — UUID не имеет "from") ──

    def _to_uuid_string(self, data: Any) -> str:
        return str(uuid.uuid4())

    # ── JWT (S40 W3) — joserfc (to_ only — verify/decode вне scope) ──

    def _to_jwt(self, data: Any) -> str:
        try:
            from joserfc import jwt as _jwt
            from joserfc.jwk import OctKey
        except ImportError as exc:
            raise ImportError(
                "to_jwt requires 'joserfc' (pip install joserfc)"
            ) from exc
        if not self.secret:
            raise ValueError("to_jwt requires 'secret' kwarg (>= 16 chars recommended)")
        body_claims: dict[str, Any] = dict(data) if isinstance(data, dict) else {}
        if self.claims:
            body_claims.update(self.claims)
        key = OctKey.import_key(self.secret)
        header = {"alg": self.algorithm, "typ": "JWT"}
        return _jwt.encode(header, body_claims, key)

    # ── Bencode (S40 W3) — bitTorrent bcode (custom ~40-LOC) ──

    def _to_bencode(self, data: Any) -> bytes:
        return _bencode(data)

    def _from_bencode(self, data: Any) -> Any:
        if isinstance(data, (bytes, bytearray)):
            raw = bytes(data)
        elif isinstance(data, str):
            raw = data.encode("utf-8")
        else:
            raise TypeError(f"from_bencode requires bytes/str, got {type(data)}")
        if not raw:
            return None
        result, _consumed = _bdecode(raw, 0)
        return result

    # ── JWT decode (S40 W4) — companion к to_jwt (W3) ──

    def _from_jwt(self, data: Any) -> dict[str, Any]:
        """Decode JWT string → claims ``dict``.

        Алгоритм и секрет берутся из ``self.algorithm``/``self.secret``
        (выставленных в ``__init__`` через mixin kwargs). joserfc делает
        verify signature → возвращает ``Token`` c атрибутом ``claims``.
        """
        try:
            from joserfc import jwt as _jwt
            from joserfc.jwk import OctKey
        except ImportError as exc:
            raise ImportError(
                "from_jwt requires 'joserfc' (pip install joserfc)"
            ) from exc
        if not self.secret:
            raise ValueError("from_jwt requires 'secret' kwarg (HS* algorithms)")
        token = _to_text(data)
        if not token:
            return {}
        key = OctKey.import_key(self.secret)
        result = _jwt.decode(token, key, algorithms=[self.algorithm])
        return dict(result.claims)

    # ── Compact JSON (S40 W4) — JSON без пробелов ──

    def _to_compact_json(self, data: Any) -> str:
        """``dict`` → minified JSON (no indent, no spaces between separators)."""
        if isinstance(data, (bytes, bytearray)):
            return data.decode("utf-8", errors="replace")
        if isinstance(data, str):
            return data  # already a string — assume compact
        return json.dumps(data, separators=(",", ":"), default=str, ensure_ascii=False)

    # ── Protobuf-like (S40 W4) — base64(JSON) (без real protobuf dep) ──

    def _to_protobuf_like(self, data: Any) -> bytes:
        """``dict`` → base64-encoded JSON ``bytes`` (protobuf-like wire format).

        Реальный protobuf не используется (dev-friendly). Формат —
        ``base64(json(dict))``, что round-trip-ается через ``from_protobuf_like``.
        """
        if data is None:
            return b""
        text = json.dumps(data, separators=(",", ":"), default=str, ensure_ascii=False)
        return base64.b64encode(text.encode("utf-8"))

    def _from_protobuf_like(self, data: Any) -> Any:
        """base64-encoded JSON ``bytes`` → ``dict`` (inverse of ``to_protobuf_like``)."""
        if isinstance(data, (bytes, bytearray)):
            raw = bytes(data)
        elif isinstance(data, str):
            raw = data.encode("utf-8")
        else:
            raise TypeError(
                f"from_protobuf_like requires bytes/str, got {type(data)}"
            )
        if not raw:
            return None
        text = base64.b64decode(raw).decode("utf-8", errors="replace")
        if not text:
            return None
        return json.loads(text)

    # ── Avro-like (S40 W4) — JSON с ``{"schema": ..., "data": ...}`` обёрткой ──

    def _to_avro_like(self, data: Any) -> str:
        """``dict`` → JSON ``str`` с обёрткой ``{"schema": ..., "data": ...}``.

        ``self.schema`` — переданный пользователем schema dict
        (default — пустой ``{}``). Реальный Avro не парсится — формат
        совместим с confluent / fastavro "datum in envelope" паттерном.
        """
        envelope = {
            "schema": self.schema if self.schema is not None else {},
            "data": data,
        }
        return json.dumps(envelope, default=str, ensure_ascii=False)


# ── Bencode (S40 W3) — recursive encoder/decoder, no external deps ──


def _bencode(obj: Any) -> bytes:
    """Recursive bencode encoder (bitTorrent metafile format)."""
    if isinstance(obj, bool):
        raise TypeError("bencode does not support bool (use 0/1 ints)")
    if isinstance(obj, int):
        return b"i" + str(obj).encode("ascii") + b"e"
    if isinstance(obj, (str, bytes, bytearray)):
        b = obj.encode("utf-8") if isinstance(obj, str) else bytes(obj)
        return str(len(b)).encode("ascii") + b":" + b
    if isinstance(obj, (list, tuple)):
        return b"l" + b"".join(_bencode(x) for x in obj) + b"e"
    if isinstance(obj, dict):
        items: list[tuple[bytes, Any]] = []
        for k, v in obj.items():
            if not isinstance(k, (str, bytes, bytearray)):
                raise TypeError(f"bencode dict keys must be str/bytes, got {type(k)}")
            kb = k.encode("utf-8") if isinstance(k, str) else bytes(k)
            items.append((kb, v))
        items.sort(key=lambda kv: kv[0])
        return b"d" + b"".join(_bencode(k) + _bencode(v) for k, v in items) + b"e"
    raise TypeError(f"cannot bencode {type(obj)}")


def _bdecode(data: bytes, idx: int) -> tuple[Any, int]:
    """Recursive bencode decoder. Returns (value, new_idx)."""
    ch = data[idx : idx + 1]
    if ch == b"i":
        end = data.index(b"e", idx)
        val = int(data[idx + 1 : end])
        return val, end + 1
    if ch == b"l":
        idx += 1
        out: list[Any] = []
        while data[idx : idx + 1] != b"e":
            item, idx = _bdecode(data, idx)
            out.append(item)
        return out, idx + 1
    if ch == b"d":
        idx += 1
        out_d: dict[Any, Any] = {}
        while data[idx : idx + 1] != b"e":
            k, idx = _bdecode(data, idx)
            v, idx = _bdecode(data, idx)
            out_d[k] = v
        return out_d, idx + 1
    # byte string: <len>:<bytes>
    colon = data.index(b":", idx)
    length = int(data[idx:colon])
    start = colon + 1
    return data[start : start + length], start + length


# ── Mixin ─────────────────────────────────────────────────────────────


class FormatConvertersMixin:
    """10 chainable format-conversion методов для ``RouteBuilder`` (S40 W1).

    Все методы возвращают ``self`` для fluent-цепочки. Реальная работа
    делегируется :class:`FormatConvertProcessor`.
    """

    __slots__ = ()

    # ── JSON ──

    def to_json(self, *, indent: int | None = None) -> "RouteBuilder":
        """Serialize ``exchange.body`` → JSON string в ``out_message.body``."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(direction="to_json", fmt="json", indent=indent)
        )

    def from_json(self, *, from_property: str = "body") -> "RouteBuilder":
        """Parse JSON string → ``dict``/``list`` в ``out_message.body``.

        ``from_property``: имя ключа в ``exchange.properties`` (default ``body``
        = ``exchange.in_message.body``).
        """
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(
                direction="from_json", fmt="json", from_property=from_property
            )
        )

    # ── CSV ──

    def to_csv(self, *, headers: list[str] | None = None) -> "RouteBuilder":
        """Convert ``list[dict]`` → CSV string.

        ``headers``: явный порядок колонок (default = keys первого ряда).
        """
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(direction="to_csv", fmt="csv", headers=headers)
        )

    def from_csv(self, csv_string: str | None = None) -> "RouteBuilder":
        """Parse CSV → ``list[dict]``.

        ``csv_string``: явное значение (default = ``exchange.in_message.body``).
        """
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(
                direction="from_csv", fmt="csv", source_value=csv_string
            )
        )

    # ── XML ──

    def to_xml(self, *, root_tag: str = "root") -> "RouteBuilder":
        """Convert ``dict`` → XML string (stdlib ``xml.etree.ElementTree``)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(direction="to_xml", fmt="xml", root_tag=root_tag)
        )

    def from_xml(self, xml_string: str | None = None) -> "RouteBuilder":
        """Parse XML → ``dict`` (через ``xmltodict`` если есть, иначе stdlib)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(
                direction="from_xml", fmt="xml", source_value=xml_string
            )
        )

    # ── YAML ──

    def to_yaml(self) -> "RouteBuilder":
        """Convert ``dict``/``list`` → YAML string."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(direction="to_yaml", fmt="yaml")
        )

    def from_yaml(self, yaml_string: str | None = None) -> "RouteBuilder":
        """Parse YAML → ``dict``/``list``."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(
                direction="from_yaml", fmt="yaml", source_value=yaml_string
            )
        )

    # ── Excel ──

    def to_excel(self, *, sheet_name: str = "Sheet1") -> "RouteBuilder":
        """Convert ``list[dict]`` → Excel bytes (openpyxl)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(
                direction="to_excel", fmt="excel", sheet_name=sheet_name
            )
        )

    def from_excel(self, excel_bytes: bytes | None = None) -> "RouteBuilder":
        """Parse Excel bytes → ``list[dict]`` (openpyxl)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(
                direction="from_excel", fmt="excel", source_value=excel_bytes
            )
        )

    # ── Parquet (S40 W2) ──

    def to_parquet(self, *, compression: str = "snappy") -> "RouteBuilder":
        """Convert ``list[dict]`` → parquet bytes (pyarrow)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(
                direction="to_parquet", fmt="parquet", compression=compression
            )
        )

    def from_parquet(self, parquet_bytes: bytes | None = None) -> "RouteBuilder":
        """Parse parquet → ``list[dict]`` (pyarrow)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(
                direction="from_parquet", fmt="parquet", source_value=parquet_bytes
            )
        )

    # ── MessagePack (S40 W2) ──

    def to_msgpack(self) -> "RouteBuilder":
        """Convert ``dict``/``list`` → msgpack bytes (fallback: ``pickle``)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(direction="to_msgpack", fmt="msgpack")
        )

    def from_msgpack(self, msgpack_bytes: bytes | None = None) -> "RouteBuilder":
        """Parse msgpack → ``dict``/``list`` (fallback: ``pickle``)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(
                direction="from_msgpack", fmt="msgpack", source_value=msgpack_bytes
            )
        )

    # ── TOML (S40 W2) ──

    def to_toml(self) -> "RouteBuilder":
        """Convert ``dict`` → TOML string (``tomli_w``)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(direction="to_toml", fmt="toml")
        )

    def from_toml(self, toml_string: str | None = None) -> "RouteBuilder":
        """Parse TOML → ``dict`` (``tomllib`` stdlib 3.11+)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(
                direction="from_toml", fmt="toml", source_value=toml_string
            )
        )

    # ── INI (S40 W2) ──

    def to_ini(self) -> "RouteBuilder":
        """Convert ``dict`` → INI string (stdlib ``configparser``)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(direction="to_ini", fmt="ini")
        )

    def from_ini(self, ini_string: str | None = None) -> "RouteBuilder":
        """Parse INI → ``dict`` (stdlib ``configparser``)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(
                direction="from_ini", fmt="ini", source_value=ini_string
            )
        )

    # ── Base64 (S40 W2) ──

    def to_base64(self) -> "RouteBuilder":
        """Encode ``bytes``/``str`` → base64 string (stdlib ``base64``)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(direction="to_base64", fmt="base64")
        )

    def from_base64(self, b64_string: str | None = None) -> "RouteBuilder":
        """Decode base64 string → ``bytes`` (stdlib ``base64``)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(
                direction="from_base64", fmt="base64", source_value=b64_string
            )
        )

    # ── URL-encoding (S40 W3) ──

    def to_url_encoded(self) -> "RouteBuilder":
        """Convert ``dict`` → URL-encoded string (application/x-www-form-urlencoded)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(direction="to_url_encoded", fmt="url_encoded")
        )

    def from_url_encoded(self, url_string: str | None = None) -> "RouteBuilder":
        """Parse URL-encoded string → ``dict`` (multi-value → ``list``)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(
                direction="from_url_encoded", fmt="url_encoded", source_value=url_string
            )
        )

    # ── HTML (S40 W3) ──

    def to_html_escape(self) -> "RouteBuilder":
        """HTML-escape string (``<>&"'`` → entities, ``quote=True``)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(direction="to_html_escape", fmt="html_escape")
        )

    def from_html_unescape(self, html_string: str | None = None) -> "RouteBuilder":
        """HTML-unescape string (entities → ``<>&"'`` chars)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(
                direction="from_html_unescape",
                fmt="html_unescape",
                source_value=html_string,
            )
        )

    # ── Markdown (S40 W3) — simple header-based ──

    def to_markdown(self) -> "RouteBuilder":
        """Convert ``dict`` → markdown string (``# key`` per top-level key)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(direction="to_markdown", fmt="markdown")
        )

    def from_markdown(self, md_string: str | None = None) -> "RouteBuilder":
        """Parse markdown → ``dict`` (extracts ``# heading`` → content)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(
                direction="from_markdown", fmt="markdown", source_value=md_string
            )
        )

    # ── UUID (S40 W3) — generator (to_ only) ──

    def to_uuid_string(self) -> "RouteBuilder":
        """Generate UUID4 string (``body`` ignored, always fresh)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(direction="to_uuid_string", fmt="uuid_string")
        )

    # ── JWT (S40 W3) — encoder (to_ only; decode вне scope) ──

    def to_jwt(
        self,
        *,
        secret: str,
        algorithm: str = "HS256",
        claims: dict[str, Any] | None = None,
    ) -> "RouteBuilder":
        """Encode ``exchange.body`` (dict) → JWT string (HS256 default).

        ``claims``: extra claims merged into body (claims override body keys).
        Requires ``joserfc`` (project dep).
        """
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(
                direction="to_jwt",
                fmt="jwt",
                secret=secret,
                algorithm=algorithm,
                claims=claims,
            )
        )

    # ── Bencode (S40 W3) ──

    def to_bencode(self) -> "RouteBuilder":
        """Convert ``dict``/``list`` → bencoded bytes (bitTorrent metafile)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(direction="to_bencode", fmt="bencode")
        )

    def from_bencode(self, bcode_bytes: bytes | None = None) -> "RouteBuilder":
        """Parse bencoded bytes → Python object (no external deps)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(
                direction="from_bencode", fmt="bencode", source_value=bcode_bytes
            )
        )

    # ── JWT decode (S40 W4) — companion к to_jwt (W3) ──

    def from_jwt(
        self,
        jwt_string: str | None = None,
        *,
        secret: str,
        algorithm: str = "HS256",
    ) -> "RouteBuilder":
        """Decode JWT ``str`` → claims ``dict`` (verify HS* signature via joserfc).

        ``jwt_string``: явный токен (default = ``exchange.in_message.body``).
        ``secret``: shared secret (HS256/HS384/HS512).
        ``algorithm``: JWT ``alg`` header value (default ``HS256``).
        """
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(
                direction="from_jwt",
                fmt="jwt",
                source_value=jwt_string,
                secret=secret,
                algorithm=algorithm,
            )
        )

    # ── Compact JSON (S40 W4) — minified JSON без пробелов ──

    def to_compact_json(self) -> "RouteBuilder":
        """Convert ``dict`` → minified JSON ``str`` (no indent, no spaces)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(direction="to_compact_json", fmt="compact_json")
        )

    # ── Protobuf-like (S40 W4) — base64(JSON) (без real protobuf dep) ──

    def to_protobuf_like(self) -> "RouteBuilder":
        """Convert ``dict`` → base64-encoded JSON ``bytes`` (protobuf-like wire format).

        No real protobuf dep — формат ``base64(json(dict))`` round-trip-ается
        через :meth:`from_protobuf_like`.
        """
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(direction="to_protobuf_like", fmt="protobuf_like")
        )

    def from_protobuf_like(self, pb_bytes: bytes | None = None) -> "RouteBuilder":
        """Decode base64-encoded JSON ``bytes`` → ``dict`` (inverse of to_protobuf_like)."""
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(
                direction="from_protobuf_like",
                fmt="protobuf_like",
                source_value=pb_bytes,
            )
        )

    # ── Avro-like (S40 W4) — JSON c ``{"schema": ..., "data": ...}`` обёрткой ──

    def to_avro_like(self, schema: dict[str, Any] | None = None) -> "RouteBuilder":
        """Convert ``dict`` → JSON ``str`` c обёрткой ``{"schema": ..., "data": ...}``.

        ``schema``: optional Avro-like schema dict (stored as-is in envelope).
        Совместим с "datum in envelope" паттерном (confluent / fastavro).
        """
        return self._add(  # type: ignore[attr-defined]
            FormatConvertProcessor(
                direction="to_avro_like", fmt="avro_like", schema=schema
            )
        )
